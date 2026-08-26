"""Turns the same dict shape `case_service.export_accepted_queries()` /
`board_service.export_test_cases()` return into a spreadsheet跑测团队能
直接打开填写的 .xlsx——JSON 导出对着代码消费方够用，但真的要交给人工去
跑，业务方一直用的是 Excel（专病管家跑测方案811.xlsx 本身就是这个格式），
不该逼着他们自己转格式。

列结构和多轮对话的 "R1/R2..." 写法，对照的是业务方那份 xlsx 里
「已设计测试用例」sheet 的真实列（用例｜Journey-场景及方向｜病例与测试
背景｜统一候选用户画像｜对应实际Query｜用户行为与对话逻辑｜测试时发送
图片），不是我们自己发明的格式——字段名不完全一样（我们多拆了几列方便
筛选），但内容对得上。

一行 = 一条用例的一套画像脚本（persona_variant）；一条用例挑了几套/带了
几套候选，就展开成几行，case/裂点/query 级别的字段每行重复。没有
persona_variants 的用例（多轮画像脚本上线之前生成的老用例）仍然单独
输出一行，画像相关列留空，query 列回退到 query.text 本身。
"""
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADERS = [
    "病例编号", "旅程阶段", "来源", "场景代码", "场景名称",
    "测试角度", "测试背景（仅供评分参考，不发给被测产品）",
    "随 query 发送的图片(seq)", "图片说明", "有标准卡",
    "预期答题要点", "红线关注点",
    "画像代码", "画像名称", "该用例下的具体表现",
    "多轮对话（R1/R2...）", "行为逻辑（考验什么）", "是否人工选用",
]

_WRAP = Alignment(wrap_text=True, vertical="top")
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _format_turns(turns: list[dict]) -> str:
    """Rn 格式，对照业务方 811 xlsx「已设计测试用例」sheet 里
    "R1（图片发送后连续发送）：\\n消息1\\n消息2" 的写法。"""
    blocks = []
    for t in turns:
        note = f"（{t.get('note')}）" if t.get("note") else ""
        header = f"R{t.get('round')}{note}："
        messages = t.get("messages") or []
        blocks.append(header + "\n" + "\n".join(messages))
    return "\n\n".join(blocks)


def build_test_case_workbook(test_cases: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    ws.freeze_panes = "A2"

    row_idx = 2
    for tc in test_cases:
        base = [
            tc.get("case_no"),
            tc.get("journey_stage"),
            "真实" if tc.get("provenance") == "real" else "推测",
            tc.get("scenario_type"),
            tc.get("scenario_name"),
            tc.get("test_direction"),
            tc.get("test_background"),
            "、".join(str(img.get("seq")) for img in (tc.get("test_images") or [])),
            tc.get("test_image_note"),
            "是" if tc.get("has_standard_card") else "否",
            "\n".join(tc.get("expected_answer_points") or []),
            "\n".join(tc.get("red_line_watch") or []),
        ]
        variants = tc.get("persona_variants") or []
        if not variants:
            # 老格式用例（多轮画像脚本上线前生成）：没有画像脚本，query 原文
            # 是唯一的实际内容，放进"多轮对话"列而不是留空——不然这一整行
            # 除了场景和答题要点，看起来什么都没有。
            row = base + [None, None, None, tc.get("query"), None, None]
            for col, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col, value=value).alignment = _WRAP
            row_idx += 1
            continue
        for v in variants:
            row = base + [
                v.get("persona_code"),
                v.get("persona_name"),
                v.get("persona_note"),
                _format_turns(v.get("turns") or []),
                v.get("behavior_logic"),
                "是" if v.get("selected") else "",
            ]
            for col, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col, value=value).alignment = _WRAP
            row_idx += 1

    widths = [16, 8, 8, 10, 16, 26, 30, 14, 20, 8, 30, 24, 12, 16, 26, 42, 30, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
