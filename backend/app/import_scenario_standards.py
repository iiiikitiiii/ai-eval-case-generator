"""One-off import of doc/专病管家测评标准-场景清单+标准.xlsx into the DB.

Replaces the phase-1 placeholder scenario_types (7 invented rows) with the
real 49 scenarios, and seeds the global evaluation rubric (20 criteria),
the 11 case-level red lines with their legal basis, and the one worked
standard-card example the source file provides.

Idempotent on re-run: scenario_types matched by scenario_number,
eval_criteria/legal_basis_refs/red_lines by their code/seq, standard_cards
by scenario_type_id — existing rows get updated in place, nothing is
duplicated.

Run after migration 0004:
    python -m app.import_scenario_standards [path/to/xlsx]
Defaults to doc/专病管家测评标准-场景清单+标准.xlsx relative to the repo root.
"""
import sys
from pathlib import Path

import openpyxl

from app.db.models.agent import ScenarioType
from app.db.models.standard import EvalCriterion, LegalBasisRef, RedLine, StandardCard, StandardCardCriterion
from app.db.session import SessionLocal

DEFAULT_PATH = Path(__file__).resolve().parents[2] / "doc" / "专病管家测评标准-场景清单+标准.xlsx"


def _bullets(cell_value: str | None) -> list[str]:
    if not cell_value:
        return []
    return [line.strip().lstrip("•").strip() for line in cell_value.split("\n") if line.strip()]


def import_scenarios(db, wb) -> int:
    """已弃用，留着只是为了不丢历史记录——它读的是旧的 8 阶段 sheet
    （"整合场景清单"），journey_stages 打的是 J01-J08。真实的六阶段模型
    见 import_scenarios_six_stage；旧的 8 阶段划分（J01 筛查与风险咨询…
    J08 复发或进展后再决策）从未出现在业务方任何文档里，是早期一版
    重建时凭空设计的，不是业务方给的分期。"""
    ws = wb["整合场景清单"]
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        journey, scenario_number, source, name, description, volume = row[:6]
        if scenario_number is None:
            continue
        stage_code = str(journey).split("-", 1)[0].strip()  # "J01-筛查与风险咨询" -> "J01"
        existing = db.query(ScenarioType).filter(ScenarioType.scenario_number == scenario_number).first()
        fields = dict(
            name=name,
            axis="patient",
            journey_stages=[stage_code],
            description=description,
            source=source,
            consultation_volume=int(volume) if isinstance(volume, (int, float)) else None,
            active=True,
        )
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(ScenarioType(code=f"SCN{int(scenario_number):02d}", scenario_number=int(scenario_number), **fields))
        n += 1
    db.commit()
    return n


# 业务方真实的六阶段旅程模型——"整合场景清单 (六阶段)" sheet。跟"整合场景
# 清单"是同一批 49 个场景（scenario_number 一一对应，用它匹配，不是重新
# 编号），只是 Patient Journey 那一列换成了业务方认可的六阶段分组：
# J01 疑诊/初筛期（合并旧 J01-J03）、J02 确诊后治疗方案决策期（合并旧
# J04-J05）、J03 初诊治疗实施期（对应旧 J06）、J04 复发/进展/耐药后治疗
# 方案调整（主要来自旧 J08）、J05 康复随访期（主要来自旧 J07）、
# J06 姑息照护期（全新——旧 J07/J08 里专门讲舒缓照护/临终关怀的场景被
# 单独拆到这里，不是机械合并）。这份映射是从两个 sheet 按 scenario_number
# 逐条比对算出来的，不是猜的。
SIX_STAGE_CODES = ["J01", "J02", "J03", "J04", "J05", "J06"]
SIX_STAGE_LABELS = {
    "J01": "疑诊 / 初筛期",
    "J02": "确诊后治疗方案决策期",
    "J03": "初诊治疗实施期",
    "J04": "复发 / 进展 / 耐药后治疗方案调整",
    "J05": "康复随访期",
    "J06": "姑息照护期",
}


def import_scenarios_six_stage(db, wb) -> int:
    """真实来源，从 0009 迁移这次落地开始用这个，不是 import_scenarios。
    幂等，按 scenario_number 匹配更新——不会重复插入，也不会因为改用这个
    sheet 就把 code=SCNxx 的编号打乱。"""
    ws = wb["整合场景清单 (六阶段)"]
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        journey, scenario_number, source, name, description, volume = row[:6]
        if scenario_number is None:
            continue
        stage_code = str(journey).split("-", 1)[0].strip()  # "J01-疑诊 / 初筛期" -> "J01"
        existing = db.query(ScenarioType).filter(ScenarioType.scenario_number == scenario_number).first()
        fields = dict(
            name=name,
            axis="patient",
            journey_stages=[stage_code],
            description=description,
            source=source,
            consultation_volume=int(volume) if isinstance(volume, (int, float)) else None,
            active=True,
        )
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(ScenarioType(code=f"SCN{int(scenario_number):02d}", scenario_number=int(scenario_number), **fields))
        n += 1
    db.commit()
    return n


def import_criteria(db, wb) -> int:
    ws = wb["标准与红线清单"]
    weights = {}
    for row in ws.iter_rows(min_row=6, max_row=9, values_only=True):
        if row[0] and row[1] is not None:
            weights[row[0]] = float(row[1])

    n = 0
    for row in ws.iter_rows(min_row=13, max_row=32, values_only=True):
        category, name, code, definition, boundary, max_points, version = row[:7]
        if not code:
            continue
        existing = db.query(EvalCriterion).filter(EvalCriterion.code == code).first()
        fields = dict(
            category=category,
            category_weight=weights.get(category, 0),
            name=name,
            definition=definition,
            evaluation_boundary=boundary,
            max_points=int(max_points),
            version=version,
        )
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(EvalCriterion(code=code, **fields))
        n += 1
    db.commit()
    return n


def import_legal_basis(db, wb) -> int:
    ws = wb["标准与红线清单"]
    n = 0
    for row in ws.iter_rows(min_row=51, max_row=57, values_only=True):
        code, title, articles, key_points, usage_note, source_url, nature = row[:7]
        if not code:
            continue
        existing = db.query(LegalBasisRef).filter(LegalBasisRef.code == code).first()
        fields = dict(title=title, articles=articles, key_points=key_points, usage_note=usage_note, source_url=source_url, nature=nature)
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(LegalBasisRef(code=code, **fields))
        n += 1
    db.commit()
    return n


def import_red_lines(db, wb) -> int:
    ws = wb["标准与红线清单"]
    n = 0
    for row in ws.iter_rows(min_row=37, max_row=47, values_only=True):
        category, seq, name, judgment, evidence, basis, verdict = row[:7]
        if seq is None:
            continue
        codes = [c.strip() for c in str(basis).split("、") if c.strip()] if basis else []
        existing = db.query(RedLine).filter(RedLine.seq == seq).first()
        fields = dict(category=category, name=name, judgment_criteria=judgment, evidence_requirements=evidence, legal_basis_codes=codes, verdict_rule=verdict)
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(RedLine(seq=int(seq), **fields))
        n += 1
    db.commit()
    return n


def import_standard_card(db, wb) -> int:
    ws = wb["标准卡示例"]
    scenario_name = ws.cell(row=4, column=3).value
    scenario_type = db.query(ScenarioType).filter(ScenarioType.name == scenario_name).first()
    if scenario_type is None:
        print(f"WARNING: 标准卡引用的场景「{scenario_name}」在场景清单里没找到，跳过标准卡导入")
        return 0

    all_red_line_seqs = [r[0] for r in ws.iter_rows(min_row=46, max_row=57, values_only=True) if isinstance(r[0], int)]

    fields = dict(
        version="V1.4",
        patient_need=ws.cell(row=7, column=3).value,
        evaluation_purpose=ws.cell(row=8, column=3).value,
        observation_conditions=ws.cell(row=9, column=3).value,
        whats_right=_bullets(ws.cell(row=13, column=1).value),
        whats_wrong=_bullets(ws.cell(row=13, column=7).value),
        applicable_red_line_seqs=all_red_line_seqs,
    )

    card = db.query(StandardCard).filter(StandardCard.scenario_type_id == scenario_type.id).first()
    if card:
        for k, v in fields.items():
            setattr(card, k, v)
        db.query(StandardCardCriterion).filter(StandardCardCriterion.standard_card_id == card.id).delete()
    else:
        card = StandardCard(scenario_type_id=scenario_type.id, **fields)
        db.add(card)
    db.commit()
    db.refresh(card)

    n = 0
    for row in ws.iter_rows(min_row=22, max_row=41, values_only=True):
        _, _, code, _, tier_a, tier_b, tier_c, tier_d, tier_e = row[:9]
        if not code:
            continue
        db.add(
            StandardCardCriterion(
                standard_card_id=card.id,
                criterion_code=code,
                tiers={"A": tier_a, "B": tier_b, "C": tier_c, "D": tier_d, "E": tier_e},
            )
        )
        n += 1
    db.commit()
    return n


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    if not path.exists():
        raise SystemExit(f"文件不存在：{path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    db = SessionLocal()
    try:
        print(f"场景清单（六阶段）：{import_scenarios_six_stage(db, wb)} 条")
        print(f"评分标准：{import_criteria(db, wb)} 条")
        print(f"法规依据：{import_legal_basis(db, wb)} 条")
        print(f"通用红线：{import_red_lines(db, wb)} 条")
        print(f"标准卡评分档：{import_standard_card(db, wb)} 条")
    finally:
        db.close()


if __name__ == "__main__":
    main()
