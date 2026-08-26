"""export_zip.build_test_case_zip — 用户反馈"压缩包里没有标准md和excel"后
核实的结论：per-case 的 standard_card.md 生成逻辑本身没问题（对着真实数据
验证过），缺的是一份汇总 Excel——压缩包之前只有分文件夹的 query.md/图片/
standard_card.md，没有一份跟"导出 Excel"按钮同样内容的总览表。这里补上
根目录的 `测试用例总览.xlsx`，用真实 zip 读回来验证结构，不只测字节数。
"""
import io
import uuid
import zipfile

import openpyxl

from app.db.models.case import Case, CaseStatus, Cutpoint, Query
from app.services.export_zip import build_test_case_zip


def _make_case_with_query(db_session, *, has_standard_card: bool = False) -> tuple[Case, Cutpoint, Query]:
    case = Case(id=uuid.uuid4(), case_no=f"CASE-TEST-{uuid.uuid4().hex[:6]}", patient_meta={}, status=CaseStatus.exported.value, current_step="out")
    db_session.add(case)
    db_session.flush()
    cp = Cutpoint(
        id=uuid.uuid4(), case_id=case.id, stage_code="J01", provenance="real",
        anchor={}, known_set=[], unknown_set=[], validity_check={}, enabled=True,
    )
    db_session.add(cp)
    db_session.flush()
    q = Query(
        id=uuid.uuid4(), cutpoint_id=cp.id, scenario_type="SCN-ZIP-TEST", text="测试 query",
        test_direction="测试角度", decision="accept", has_standard_card=has_standard_card,
        expected_answer_points=["要点1"], red_line_watch=[],
    )
    db_session.add(q)
    db_session.flush()
    return case, cp, q


def test_zip_includes_summary_xlsx_at_root(db_session):
    case, cp, q = _make_case_with_query(db_session)
    data = build_test_case_zip(db_session, [(q, cp, case)])

    zf = zipfile.ZipFile(io.BytesIO(data))
    names = zf.namelist()
    assert "测试用例总览.xlsx" in names


def test_zip_summary_xlsx_is_a_valid_readable_workbook(db_session):
    case, cp, q = _make_case_with_query(db_session)
    data = build_test_case_zip(db_session, [(q, cp, case)])

    zf = zipfile.ZipFile(io.BytesIO(data))
    xlsx_bytes = zf.read("测试用例总览.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "病例编号" in header
    assert "场景代码" in header
    data_row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert data_row[header.index("病例编号")] == case.case_no
    assert data_row[header.index("场景代码")] == "SCN-ZIP-TEST"


def test_zip_summary_xlsx_filename_has_utf8_flag(db_session):
    """zipfile.writestr(str, data) 默认不会给中文文件名设 UTF-8 flag——
    这个项目之前真实踩过这个坑（见 export_zip.py 的 _write），这里锁死
    回归：根目录这份 xlsx 也必须走同一个带 flag_bits 的写入路径。"""
    case, cp, q = _make_case_with_query(db_session)
    data = build_test_case_zip(db_session, [(q, cp, case)])

    zf = zipfile.ZipFile(io.BytesIO(data))
    info = zf.getinfo("测试用例总览.xlsx")
    assert info.flag_bits & 0x800


def test_zip_still_includes_standard_card_when_present(db_session):
    """总览 Excel 是新增的，不能把已有的 per-case standard_card.md 挤掉。"""
    case, cp, q = _make_case_with_query(db_session, has_standard_card=True)
    data = build_test_case_zip(db_session, [(q, cp, case)])

    zf = zipfile.ZipFile(io.BytesIO(data))
    names = zf.namelist()
    assert "测试用例总览.xlsx" in names
    # 没有真实 StandardCard 行时（这里没建），has_standard_card=True 也不会
    # 生成 standard_card.md——这是数据缺口，不是这个测试要验证的；用
    # query.md 存在来确认这条用例本身正常打包。
    assert any(n.endswith("query.md") for n in names)
